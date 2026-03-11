"""
Manage resume file storage and Excel tracking.
- Saves resumes to resume/{company_name}/
- Maintains tracker.xlsx with columns for each resume version
"""

import os
import shutil
from datetime import datetime

from openpyxl import Workbook, load_workbook


RESUME_DIR = "resume"
TRACKER_FILE = "tracker.xlsx"


def sanitize_company_name(name: str) -> str:
    """Clean company name for use as directory name."""
    # Remove / replace characters that are problematic in paths
    cleaned = name.strip()
    for ch in r'<>:"/\|?*':
        cleaned = cleaned.replace(ch, "_")
    # Collapse multiple underscores/spaces
    cleaned = "_".join(cleaned.split())
    return cleaned


def company_exists(company: str, base_dir: str) -> bool:
    """Check if a company folder already exists in the resume directory."""
    company_dir = os.path.join(base_dir, RESUME_DIR, sanitize_company_name(company))
    return os.path.isdir(company_dir)


def save_resume(
    company: str,
    tex_content: str,
    pdf_path: str,
    base_dir: str,
    is_second: bool = False,
) -> dict:
    """
    Save resume files to resume/{company}/ and return paths.
    is_second=True means this is the second resume (resume2 variant).
    """
    company_safe = sanitize_company_name(company)
    company_dir = os.path.join(base_dir, RESUME_DIR, company_safe)
    os.makedirs(company_dir, exist_ok=True)

    suffix = "_v2" if is_second else ""
    tex_name = f"resume{suffix}.tex"
    pdf_name = f"resume{suffix}.pdf"

    tex_dest = os.path.join(company_dir, tex_name)
    pdf_dest = os.path.join(company_dir, pdf_name)

    # Save .tex
    with open(tex_dest, "w", encoding="utf-8") as f:
        f.write(tex_content)

    # Copy .pdf
    shutil.copy2(pdf_path, pdf_dest)

    print(f"[+] Saved: {tex_dest}")
    print(f"[+] Saved: {pdf_dest}")

    return {
        "tex_path": tex_dest,
        "pdf_path": pdf_dest,
        "company": company,
        "company_dir": company_dir,
    }


def update_tracker(
    company: str,
    jd_url: str,
    pdf_path: str,
    base_dir: str,
    is_second: bool = False,
):
    """
    Update tracker.xlsx:
    - Columns: Company, JD URL, Resume Path, Date, Resume2 Path (if applicable)
    - If same company already exists and is_second, add Resume2 Path column entry.
    """
    tracker_path = os.path.join(base_dir, TRACKER_FILE)
    headers = ["Company", "JD URL", "Resume Path", "Date", "Resume2 Path"]

    if os.path.exists(tracker_path):
        wb = load_workbook(tracker_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Resume Tracker"
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

    # Ensure Resume2 Path column header exists
    if ws.cell(row=1, column=5).value != "Resume2 Path":
        ws.cell(row=1, column=5, value="Resume2 Path")

    # Make path relative to base_dir for cleaner tracking
    rel_path = os.path.relpath(pdf_path, base_dir)
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    if is_second:
        # Find existing row for this company and add Resume2 Path
        found = False
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == company:
                ws.cell(row=row, column=5, value=rel_path)
                found = True
                break
        if not found:
            # Company row doesn't exist yet, add new with resume2 col
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1, value=company)
            ws.cell(row=next_row, column=2, value=jd_url)
            ws.cell(row=next_row, column=4, value=now)
            ws.cell(row=next_row, column=5, value=rel_path)
    else:
        # Add new row
        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1, value=company)
        ws.cell(row=next_row, column=2, value=jd_url)
        ws.cell(row=next_row, column=3, value=rel_path)
        ws.cell(row=next_row, column=4, value=now)

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(tracker_path)
    print(f"[+] Updated tracker: {tracker_path}")
