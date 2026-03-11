"""
AutoJD - Automatically tailor your resume to any job description.

Usage:
    python -m autojd <JD_URL> <COMPANY_NAME> [--api-key KEY]

Environment:
    OPENAI_API_KEY - OpenAI API key (or pass via --api-key)
"""

import argparse
import os
import re
import sys
from urllib.parse import urlparse
from openpyxl import load_workbook
from dotenv import load_dotenv

from autojd.fetcher import fetch_jd
from autojd.modifier import modify_resume
from autojd.pdf_builder import compile_and_fit, ensure_latex_available
from autojd.storage import company_exists, sanitize_company_name, save_resume, update_tracker


BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Load variables from root .env (if present), without overriding existing env vars.
load_dotenv(os.path.join(BASE_DIR, ".env"), override=False)


class SkipJobError(RuntimeError):
    """Raised when a job should be skipped with a user-facing reason."""


def _hostname_first_label(url: str) -> str:
    host = urlparse(url).hostname or ""
    if not host:
        return "unknown"
    return host.split(".")[0] or "unknown"


def _normalize_source(source_line: str) -> str:
    low = source_line.lower()
    if "workday" in low:
        return "workday"
    if "smartrecruiters" in low:
        return "smartrecruiters"
    if "oracle" in low:
        return "oracle"
    if "lever" in low:
        return "lever"
    if "eightfold" in low:
        return "eightfold"
    if "greenhouse" in low:
        return "greenhouse"
    return "unknown"


def _source_from_url(url: str) -> str:
    host = (urlparse(url).hostname or "").lower()
    if "myworkdayjobs.com" in host:
        return "workday"
    if "smartrecruiters.com" in host:
        return "smartrecruiters"
    if "oraclecloud.com" in host:
        return "oracle"
    if "lever.co" in host:
        return "lever"
    if "eightfold.ai" in host:
        return "eightfold"
    if "greenhouse.io" in host:
        return "greenhouse"
    return "unknown"


def _company_from_url(url: str, source: str) -> str:
    parsed = urlparse(url)
    host = parsed.hostname or ""

    # User rule: Workday company is first word after https://
    if source == "workday":
        return _hostname_first_label(url)

    # User rule: Greenhouse company is first path segment after greenhouse.io
    if source == "greenhouse":
        parts = [p for p in (parsed.path or "").split("/") if p]
        # Typical greenhouse URL: /<company>/jobs/<id>
        if parts:
            return parts[0]
        return _hostname_first_label(url)

    # Existing behavior for other sources.
    if source in {"eightfold", "oracle"}:
        return _hostname_first_label(url)

    return _hostname_first_label(url)


def _parse_scraper_jobs(text: str) -> list[dict[str, str]]:
    """Parse scraper feed text or plain URL list into normalized job entries."""
    entry_pattern = re.compile(
        r"(?ms)^\s*(?P<source_line>[^\n]*NEW JOB[^\n]*)\n"
        r"Title:\s*(?P<title>[^\n]+)\n"
        r"(?:Company:\s*(?P<company>[^\n]+)\n)?"
        r"Location:\s*(?P<location>[^\n]+)\n"
        r"Link:\s*(?P<link>https?://\S+)",
    )

    raw_jobs: list[dict[str, str]] = []
    for m in entry_pattern.finditer(text):
        source = _normalize_source(m.group("source_line") or "")
        title = (m.group("title") or "").strip()
        company = (m.group("company") or "").strip()
        location = (m.group("location") or "").strip()
        link = (m.group("link") or "").strip()

        company_missing = (not company) or company.lower() == "unknown"

        # User rules / inference for missing company values.
        if source in {"eightfold", "oracle", "workday", "greenhouse"}:
            company = _company_from_url(link, source)
        elif company_missing:
            company = _company_from_url(link, source)

        raw_jobs.append(
            {
                "source": source,
                "title": title,
                "company": company,
                "location": location,
                "link": link,
            }
        )

    # Fallback: plain link-only input (one URL per line or mixed text).
    if not raw_jobs:
        links = re.findall(r"https?://\S+", text)
        for link in links:
            clean_link = link.strip().rstrip(",")
            source = _source_from_url(clean_link)
            company = _company_from_url(clean_link, source)
            raw_jobs.append(
                {
                    "source": source,
                    "title": "(title unavailable)",
                    "company": company,
                    "location": "(location unavailable)",
                    "link": clean_link,
                }
            )

    # Deduplicate by canonical link (case-insensitive)
    deduped: list[dict[str, str]] = []
    seen: set[str] = set()
    for job in raw_jobs:
        key = job["link"].lower().rstrip("/")
        if key in seen:
            continue
        seen.add(key)
        deduped.append(job)
    return deduped


def _company_has_two_resumes(base_dir: str, company: str) -> bool:
    """Return True when both first and second resume variants already exist."""
    # Source of truth: tracker.xlsx with Resume Path + Resume2 Path for same company.
    tracker_path = os.path.join(base_dir, "tracker.xlsx")
    if os.path.exists(tracker_path):
        try:
            wb = load_workbook(tracker_path, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_company = (row[0] or "").strip() if len(row) > 0 and row[0] else ""
                resume_path = row[2] if len(row) > 2 else None
                resume2_path = row[4] if len(row) > 4 else None
                if row_company.lower() == company.strip().lower() and resume_path and resume2_path:
                    wb.close()
                    return True
            wb.close()
        except Exception:
            # Fall back to file-system checks below.
            pass

    # Fallback: check resume folder files.
    company_safe = sanitize_company_name(company)
    company_dir = os.path.join(base_dir, "resume", company_safe)
    if not os.path.isdir(company_dir):
        return False

    has_primary = os.path.exists(os.path.join(company_dir, "resume.pdf")) or os.path.exists(
        os.path.join(company_dir, "resume.tex")
    )
    has_second = os.path.exists(os.path.join(company_dir, "resume_v2.pdf")) or os.path.exists(
        os.path.join(company_dir, "resume_v2.tex")
    )
    return has_primary and has_second


def run_batch(batch_text: str, api_key: str | None = None, dry_run: bool = False, limit: int | None = None):
    jobs = _parse_scraper_jobs(batch_text)
    if limit is not None and limit > 0:
        jobs = jobs[:limit]

    if not jobs:
        raise RuntimeError("No jobs could be parsed from batch text.")

    print(f"[*] Parsed {len(jobs)} jobs from batch input")
    for idx, job in enumerate(jobs, 1):
        print(
            f"  [{idx}] {job['source']} | {job['company']} | {job['title']}\n"
            f"      {job['link']}"
        )

    if dry_run:
        print("[*] Dry run enabled. No resume generation executed.")
        return

    failures: list[tuple[dict[str, str], str]] = []
    skipped: list[tuple[dict[str, str], str]] = []
    successes: list[dict[str, str]] = []
    for idx, job in enumerate(jobs, 1):
        print("\n" + "=" * 60)
        print(f"[*] Batch job {idx}/{len(jobs)}: {job['company']} - {job['title']}")
        print("=" * 60)
        if _company_has_two_resumes(BASE_DIR, job["company"]):
            reason = "Company already has resume.pdf and resume_v2.pdf (or .tex variants)."
            skipped.append((job, reason))
            print(f"[*] Skipping {job['company']}: {reason}")
            continue
        try:
            run(job["link"], job["company"], api_key=api_key)
            successes.append(job)
        except SkipJobError as exc:
            skipped.append((job, str(exc)))
            print(f"[*] Skipping {job['company']}: {exc}")
        except Exception as exc:
            failures.append((job, str(exc)))
            print(f"[!] Failed for {job['link']}: {exc}")

    print("\n" + "=" * 60)
    print("[*] Batch Summary")
    print("=" * 60)
    print(f"[+] Success: {len(successes)}")
    for job in successes:
        print(f"  - {job['company']} | {job['title']}\n    {job['link']}")

    print(f"[!] Failed: {len(failures)}")
    for job, err in failures:
        print(
            f"  - {job['company']} | {job['title']}\n"
            f"    {job['link']}\n"
            f"    Reason: {err}"
        )

    print(f"[*] Skipped: {len(skipped)}")
    for job, reason in skipped:
        print(
            f"  - {job['company']} | {job['title']}\n"
            f"    {job['link']}\n"
            f"    Reason: {reason}"
        )


def get_resume_template(base_dir: str, use_second: bool) -> str:
    """Read the appropriate resume template."""
    filename = "resume2.tex" if use_second else "resume.tex"
    path = os.path.join(base_dir, filename)
    if not os.path.exists(path):
        raise FileNotFoundError(f"Resume template not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


def run(url: str, company: str, api_key: str | None = None):
    """Main pipeline: fetch JD -> modify resume -> build PDF -> save & track."""
    base_dir = BASE_DIR

    # Preflight: fail early if LaTeX compiler is missing.
    ensure_latex_available()

    # Check if company already exists (use resume2.tex in that case)
    is_second = company_exists(company, base_dir)
    if is_second:
        print(f"[*] Company '{company}' already exists. Using resume2.tex for second version.")
    else:
        print(f"[*] New company '{company}'. Using resume.tex.")

    # Step 1: Fetch JD
    jd_text = fetch_jd(url)
    print(f"\n{'='*60}")
    print("JD Preview (first 500 chars):")
    print(jd_text)
    print(f"{'='*60}\n")

    # Step 2: Read resume template
    resume_tex = get_resume_template(base_dir, is_second)

    # Step 3: Modify resume with OpenAI
    tailored_tex = modify_resume(jd_text, resume_tex, api_key=api_key)

    # Step 4: Compile to PDF and ensure 1 page
    final_tex, pdf_path = compile_and_fit(tailored_tex)

    # Step 5: Save to resume/{company}/
    result = save_resume(company, final_tex, pdf_path, base_dir, is_second=is_second)

    # Step 6: Update tracker.xlsx
    update_tracker(company, url, result["pdf_path"], base_dir, is_second=is_second)

    print(f"\n{'='*60}")
    print(f"[✓] Done! Resume saved to: {result['company_dir']}")
    print(f"{'='*60}")


def main():
    parser = argparse.ArgumentParser(
        description="AutoJD - Tailor your resume to any job description",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            '  python -m autojd "https://jobs.example.com/12345" "Google"\n'
            '  python -m autojd "https://boards.greenhouse.io/..." "Meta" --api-key sk-...\n'
        ),
    )
    parser.add_argument("url", nargs="?", help="URL of the job posting")
    parser.add_argument("company", nargs="?", help="Company name (used for folder naming)")
    parser.add_argument("--api-key", help="OpenAI API key (or set OPENAI_API_KEY env var)")
    parser.add_argument("--batch-file", help="Path to text file containing scraper job feed")
    parser.add_argument("--dry-run", action="store_true", help="Parse and preview batch jobs only")
    parser.add_argument("--limit", type=int, help="Only process first N parsed batch jobs")

    args = parser.parse_args()
    if args.batch_file:
        with open(args.batch_file, "r", encoding="utf-8") as f:
            batch_text = f.read()
        run_batch(
            batch_text=batch_text,
            api_key=args.api_key,
            dry_run=args.dry_run,
            limit=args.limit,
        )
        return

    if not args.url or not args.company:
        parser.error("Provide <url> <company> for single mode, or use --batch-file for batch mode.")

    run(args.url, args.company, api_key=args.api_key)


if __name__ == "__main__":
    main()
